import copy

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox
)

from maze.maze import Maze
from maze.cell import Cell
from maze.generator import generate_maze_prim
from maze.export import maze_to_text
from algorithms import ALGORITHMS
from simulation.simulator import simulate
from analysis.results import AlgorithmResult
from analysis.plots import plot_results

from openpyxl import Workbook


def manhattan(a, b):
    return abs(a.row - b.row) + abs(a.col - b.col)


class TestTab(QWidget):
    def __init__(self):
        super().__init__()
        self.results = []
        self.grid = None
        self.start = None
        self.goal = None

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.run_btn = QPushButton("Rozpocznij testy")
        self.run_btn.clicked.connect(self.run_tests)
        layout.addWidget(self.run_btn)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(
            ["Algorytm", "Kroki", "Czas [s]"]
        )
        layout.addWidget(self.table)

        self.save_txt_btn = QPushButton("Zapisz wyniki do TXT")
        self.save_txt_btn.clicked.connect(self.save_results_txt)
        layout.addWidget(self.save_txt_btn)

        self.export_excel_btn = QPushButton("Eksportuj wyniki do Excela")
        self.export_excel_btn.clicked.connect(self.export_results_excel)
        layout.addWidget(self.export_excel_btn)

        self.export_plot_excel_btn = QPushButton(
            "Eksportuj dane wykresu do Excela"
        )
        self.export_plot_excel_btn.clicked.connect(
            self.export_plot_data_excel
        )
        layout.addWidget(self.export_plot_excel_btn)

        self.plot_btn = QPushButton("Zobacz wykres")
        self.plot_btn.clicked.connect(self.show_plot)
        layout.addWidget(self.plot_btn)

        self.setLayout(layout)

    def run_tests(self):
        self.grid = generate_maze_prim(21, 21)
        self.start = (1, 1)
        self.goal = (19, 19)

        self.results.clear()
        self.table.setRowCount(0)

        for name, algorithm in ALGORITHMS.items():
            maze = Maze(copy.deepcopy(self.grid))
            start_cell = Cell(*self.start)
            goal_cell = Cell(*self.goal)

            explorer = algorithm(maze, start_cell, goal_cell)
            steps = simulate(explorer)

            if not steps:
                continue

            result = AlgorithmResult(
                name=name,
                steps=len(steps),
                time=steps[-1].time_sec,
                path=[s.cell for s in steps]
            )
            self.results.append(result)

            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(name))
            self.table.setItem(row, 1, QTableWidgetItem(str(result.steps)))
            self.table.setItem(row, 2, QTableWidgetItem(f"{result.time:.2f}"))

        QMessageBox.information(self, "Gotowe", "Testy zakończone")

    def save_results_txt(self):
        if not self.results:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Zapisz wyniki", "", "Plik tekstowy (*.txt)"
        )
        if not path:
            return

        with open(path, "w", encoding="utf-8") as f:
            f.write("LABIRYNT:\n")
            f.write(maze_to_text(self.grid, self.start, self.goal))
            f.write("\n\nWYNIKI:\n")

            for r in self.results:
                f.write(
                    f"{r.name}: {r.steps} kroków, {r.time:.2f} s\n"
                )

        QMessageBox.information(self, "Zapisano", "Plik TXT zapisany")

    def export_results_excel(self):
        if not self.results:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Eksport do Excela", "", "Plik Excel (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Wyniki"

        ws.append(["Algorytm", "Kroki", "Czas [s]"])

        for r in self.results:
            ws.append([r.name, r.steps, r.time])

        wb.save(path)
        QMessageBox.information(self, "Zapisano", "Excel z wynikami zapisany")

    def export_plot_data_excel(self):
        if not self.results:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Eksport danych wykresu", "", "Plik Excel (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        goal_cell = Cell(*self.goal)

        for result in self.results:
            ws = wb.create_sheet(title=result.name[:31])
            ws.append(
                ["Krok", "Dystans Manhattan", "Row", "Col"]
            )

            for i, cell in enumerate(result.path):
                ws.append([
                    i,
                    manhattan(cell, goal_cell),
                    cell.row,
                    cell.col
                ])

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(path)
        QMessageBox.information(
            self, "Zapisano", "Dane wykresu zapisane do Excela"
        )

    def show_plot(self):
        if not self.results:
            return
        goal_cell = Cell(*self.goal)
        plot_results(self.results, goal_cell)
